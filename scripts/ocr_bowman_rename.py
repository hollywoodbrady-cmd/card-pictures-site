import argparse
import csv
import re
from pathlib import Path
from difflib import SequenceMatcher

from PIL import Image, ImageOps, ImageEnhance
import pytesseract

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ----------------------------
# Text helpers
# ----------------------------

def slugify(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[’'`]", "", s)
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s


def clean_text(s: str) -> str:
    s = (s or "").replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def normalize_name(s: str) -> str:
    # normalize for matching: letters/spaces/hyphens only
    s = (s or "").strip()
    s = s.replace(",", " ")
    s = re.sub(r"[^A-Za-z \-]", " ", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    # remove common junk words
    s = re.sub(r"\b(bowman|draft|chrome|pro|debut)\b", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def normalize_team(s: str) -> str:
    s = (s or "").strip()
    s = s.replace(",", " ")
    s = re.sub(r"[^A-Za-z \-]", " ", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def best_match(query: str, candidates: list[str], min_len=4):
    q = (query or "").strip()
    if len(re.sub(r"[^A-Za-z]", "", q)) < min_len:
        return ("", 0.0)

    best = ("", 0.0)
    for c in candidates:
        sc = similarity(q, c)
        if sc > best[1]:
            best = (c, sc)

    return best  # (candidate, score)


# ----------------------------
# Checklist dictionary loading
# ----------------------------

def load_checklist_dictionary(xlsx_paths: list[str]):
    """
    Loads players + teams from checklist sheets.

    Works with your checklists because they contain rows like:
        Card# | Player | Team
    typically starting after some header text.

    Returns: (players_sorted, teams_sorted)
    """
    if openpyxl is None:
        raise RuntimeError("Missing dependency: openpyxl. Run: pip install openpyxl")

    players = set()
    teams = set()

    def add_row(player, team):
        p = normalize_name(player)
        t = normalize_team(team)
        if p and len(p) >= 4:
            # strip trailing commas some sheets include
            p = p.rstrip(",").strip()
            players.add(p)
        if t and len(t) >= 3:
            teams.add(t)

    for xp in xlsx_paths:
        wb = openpyxl.load_workbook(xp, read_only=True, data_only=True)
        # These sheets exist in your uploaded files
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() not in ("base", "autographs", "inserts", "full checklist", "teams", "team sets"):
                continue
            ws = wb[sheet_name]

            for row in ws.iter_rows(values_only=True):
                # Expect: [Card#, Player, Team, ...]
                if not row or len(row) < 3:
                    continue
                card, player, team = row[0], row[1], row[2]

                # Skip obvious header rows (card is None or "Parallels" etc.)
                if not player or not team:
                    continue
                if isinstance(player, str) and player.strip().lower() in ("parallels", "base set", "autographs", "inserts"):
                    continue

                # Card number is often like "PD-1" / "BD-14"; helps filter real rows
                if isinstance(card, str) and re.match(r"^[A-Z]{1,3}\-\d+", card.strip()):
                    add_row(player, team)
                else:
                    # Some sheets might omit card# formatting; still try but be cautious
                    if isinstance(player, str) and isinstance(team, str) and len(player) >= 4 and len(team) >= 3:
                        add_row(player, team)

    return (sorted(players), sorted(teams))


# ----------------------------
# OCR
# ----------------------------

def preprocess(img: Image.Image) -> Image.Image:
    img = ImageOps.exif_transpose(img)
    img = img.convert("RGB")
    img = ImageOps.grayscale(img)
    img = ImageEnhance.Contrast(img).enhance(2.0)
    img = ImageEnhance.Sharpness(img).enhance(1.5)
    return img


def ocr_region(img: Image.Image, box, psm=6, upscale=2) -> str:
    cropped = img.crop(box)
    if upscale and upscale > 1:
        cropped = cropped.resize((cropped.width * upscale, cropped.height * upscale))
    cfg = f"--oem 3 --psm {psm}"
    txt = pytesseract.image_to_string(cropped, config=cfg)
    return clean_text(txt)


def ocr_name_line(img: Image.Image, box, upscale=3, thresh=160, invert=False) -> str:
    cropped = img.crop(box)
    if upscale and upscale > 1:
        cropped = cropped.resize((cropped.width * upscale, cropped.height * upscale))
    cropped = ImageOps.grayscale(cropped)
    cropped = ImageOps.autocontrast(cropped)
    if invert:
        cropped = ImageOps.invert(cropped)
    cropped = cropped.point(lambda p: 255 if p > thresh else 0)

    # IMPORTANT: no stray "-" argument
    cfg = "--oem 3 --psm 7 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz-"
    txt = pytesseract.image_to_string(cropped, config=cfg)
    return clean_text(txt)


def detect_first_bowman(img: Image.Image, fast=False) -> bool:
    w, h = img.size
    box = (0, 0, int(w * 0.33), int(h * 0.33))
    txt = ocr_region(img, box, psm=6, upscale=(1 if fast else 2)).lower()
    return ("1st" in txt) or ("first" in txt and "bowman" in txt)


def extract_player_ocr(img: Image.Image, fast=False):
    w, h = img.size
    boxes = [
        (int(w * 0.12), int(h * 0.02), int(w * 0.88), int(h * 0.18)),  # top
        (int(w * 0.12), int(h * 0.10), int(w * 0.88), int(h * 0.28)),  # upper-mid
        (int(w * 0.12), int(h * 0.58), int(w * 0.88), int(h * 0.74)),  # lower-mid
        (int(w * 0.12), int(h * 0.62), int(w * 0.88), int(h * 0.80)),  # lower
    ]
    if fast:
        boxes = [boxes[0], boxes[2]]

    name_upscale = 2 if fast else 3
    thresh = 170 if fast else 160

    best_raw = ""
    best_idx = -1
    best_inv = False

    for i, box in enumerate(boxes):
        for inv in (False, True):
            raw = ocr_name_line(img, box, upscale=name_upscale, thresh=thresh, invert=inv)
            raw_norm = normalize_name(raw)
            if len(raw_norm) > len(normalize_name(best_raw)):
                best_raw = raw
                best_idx = i
                best_inv = inv

    return best_raw, best_idx, best_inv


def build_filename(player: str, team: str, set_token: str, is_first: bool) -> str:
    p = slugify(player)[:40]
    t = slugify(team)[:30] if team else ""
    sg = slugify(set_token)[:40] if set_token else ""

    parts = [p]
    if t:
        parts.append(t)
    if sg:
        parts.append(sg)
    if is_first and ("bowman" in (set_token or "")):
        parts.append("1st-bowman")
    parts.append("front")
    return "__".join(parts) + ".jpg"


# ----------------------------
# Main
# ----------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True, help="Folder containing IMG_####.JPEG files")
    ap.add_argument("--out", required=True, help="Output CSV path (rename map)")
    ap.add_argument("--tesseract", default=r"C:\Program Files\Tesseract-OCR\tesseract.exe")
    ap.add_argument("--checklist", action="append", default=[], help="Checklist .xlsx file (can repeat)")
    ap.add_argument("--set-token", default="bowman-2025-pro-debut", help="Token to include in filename")
    ap.add_argument("--min-player-score", type=float, default=0.84, help="Player fuzzy match threshold (0-1)")
    ap.add_argument("--min-team-score", type=float, default=0.80, help="Team fuzzy match threshold (0-1)")
    ap.add_argument("--max-files", type=int, default=0, help="Process only first N files (0 = all)")
    ap.add_argument("--fast", action="store_true", help="Faster OCR (less accurate)")
    ap.add_argument("--no-first-detect", action="store_true", help="Skip 1st Bowman detection (faster)")
    args = ap.parse_args()

    pytesseract.pytesseract.tesseract_cmd = args.tesseract

    players = []
    teams = []
    if args.checklist:
        players, teams = load_checklist_dictionary(args.checklist)
        print(f"Loaded dictionary: {len(players)} players, {len(teams)} teams")

    src_dir = Path(args.dir)
    out_csv = Path(args.out)

    files = sorted([p for p in src_dir.iterdir()
                    if p.is_file() and p.suffix.lower() in [".jpg", ".jpeg"]])

    if args.max_files and args.max_files > 0:
        files = files[:args.max_files]

    total = len(files)
    rows = []

    for idx, f in enumerate(files, start=1):
        print(f"[{idx}/{total}] {f.name}")

        try:
            img0 = Image.open(f)
            img = preprocess(img0)
            w, h = img.size

            # Team region near bottom (works decently on most sets)
            team_box = (int(w * 0.12), int(h * 0.76), int(w * 0.88), int(h * 0.92))

            # OCR raw
            player_raw, name_box_idx, used_invert = extract_player_ocr(img, fast=args.fast)
            team_raw = ocr_region(img, team_box, psm=6, upscale=(1 if args.fast else 2))

            player_guess = normalize_name(player_raw)
            team_guess = normalize_team(team_raw)

            # Dictionary snap (fuzzy match)
            player_match, player_score = ("", 0.0)
            team_match, team_score = ("", 0.0)

            if players and player_guess:
                player_match, player_score = best_match(player_guess, players)
            if teams and team_guess:
                team_match, team_score = best_match(team_guess, teams, min_len=3)

            # Decide final
            use_player = player_match if player_score >= args.min_player_score else ""
            use_team = team_match if team_score >= args.min_team_score else ""

            is_first = False
            if not args.no_first_detect:
                is_first = detect_first_bowman(img, fast=args.fast)

            if use_player:
                new_name = build_filename(use_player, use_team, args.set_token, is_first)
            else:
                new_name = ""  # skip rename; you can manually fix later

            rows.append({
                "OldName": f.name,
                "NewName": new_name,
                "PlayerOCR": player_guess,
                "PlayerMatch": use_player,
                "PlayerScore": round(player_score, 3),
                "TeamOCR": team_guess,
                "TeamMatch": use_team,
                "TeamScore": round(team_score, 3),
                "SetToken": args.set_token,
                "Is1stBowman": "yes" if is_first else "no",
                "NameBoxIdx": name_box_idx,
                "UsedInvert": "yes" if used_invert else "no",
                "Error": ""
            })

        except Exception as e:
            rows.append({
                "OldName": f.name,
                "NewName": "",
                "PlayerOCR": "",
                "PlayerMatch": "",
                "PlayerScore": 0,
                "TeamOCR": "",
                "TeamMatch": "",
                "TeamScore": 0,
                "SetToken": args.set_token,
                "Is1stBowman": "no",
                "NameBoxIdx": -1,
                "UsedInvert": "no",
                "Error": str(e)
            })

    out_csv.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "OldName", "NewName",
        "PlayerOCR", "PlayerMatch", "PlayerScore",
        "TeamOCR", "TeamMatch", "TeamScore",
        "SetToken", "Is1stBowman",
        "NameBoxIdx", "UsedInvert",
        "Error"
    ]

    with out_csv.open("w", newline="", encoding="utf-8") as fp:
        wri = csv.DictWriter(fp, fieldnames=fieldnames)
        wri.writeheader()
        wri.writerows(rows)

    print(f"\nWrote rename map: {out_csv}")
    print(f"Files processed: {total}")
    print("Rows with blank NewName need manual review (or lower thresholds).")


if __name__ == "__main__":
    main()
