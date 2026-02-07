#!/usr/bin/env python3
import argparse
import csv
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional

from PIL import Image, ImageOps, ImageEnhance
import pytesseract
from openpyxl import load_workbook


# ----------------------------
# Helpers
# ----------------------------
def slugify(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[’'`]", "", s)
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s


def clean_text(s: str) -> str:
    s = (s or "").replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def extract_img_num(p: Path) -> str:
    """Return digits after IMG_#### in filename, else ''."""
    m = re.search(r"IMG_(\d{4,})", p.name, flags=re.IGNORECASE)
    return m.group(1) if m else ""


def is_image_file(p: Path) -> bool:
    return p.is_file() and p.suffix.lower() in [".jpg", ".jpeg", ".png", ".webp"]


# ----------------------------
# Checklist loading
# ----------------------------
@dataclass(frozen=True)
class ChecklistEntry:
    code: str
    player: str
    team: str


def load_checklist_map(xlsx_path: Path) -> Dict[str, ChecklistEntry]:
    """
    Load card codes from the Pro Debut checklist workbook.
    Expected headers in key sheets: CardCode, Player, Team
    """
    wb = load_workbook(str(xlsx_path), data_only=True)
    code_map: Dict[str, ChecklistEntry] = {}

    # Use the sheets that actually have CardCode/Player/Team
    candidate_sheets = ["Base", "Autographs", "Inserts", "Full Checklist"]
    for sheet_name in candidate_sheets:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        headers = [(str(c.value).strip() if c.value is not None else "") for c in ws[1]]
        # Find header indexes
        def find_col(name: str) -> int:
            for i, h in enumerate(headers):
                if h.lower() == name.lower():
                    return i
            return -1

        i_code = find_col("CardCode")
        i_player = find_col("Player")
        i_team = find_col("Team")
        if i_code < 0:
            continue

        for r in ws.iter_rows(min_row=2, values_only=True):
            code = (str(r[i_code]).strip() if i_code < len(r) and r[i_code] is not None else "")
            if not code:
                continue
            code = code.upper()

            player = (str(r[i_player]).strip() if i_player >= 0 and i_player < len(r) and r[i_player] is not None else "")
            team = (str(r[i_team]).strip() if i_team >= 0 and i_team < len(r) and r[i_team] is not None else "")

            # Prefer first seen; you can change this if you want last-wins
            if code not in code_map:
                code_map[code] = ChecklistEntry(code=code, player=player, team=team)

    return code_map


# ----------------------------
# Pairing (AUTO OFFSET)
# ----------------------------
def pair_front_back_auto(front_dir: Path, back_dir: Path, max_offset: int = 10) -> Tuple[int, List[Tuple[Path, Path]]]:
    """
    Pair by IMG_#### number with auto offset detection.

    If offset = +1, it means:
        back_num = front_num + 1

    Returns (best_offset, pairs)
    """
    fronts: Dict[int, Path] = {}
    backs: Dict[int, Path] = {}

    for p in front_dir.iterdir():
        if is_image_file(p):
            n = extract_img_num(p)
            if n:
                fronts[int(n)] = p

    for p in back_dir.iterdir():
        if is_image_file(p):
            n = extract_img_num(p)
            if n:
                backs[int(n)] = p

    if not fronts or not backs:
        return 0, []

    # Try offsets that maximize matches
    best_offset = 0
    best_count = -1

    for off in range(-max_offset, max_offset + 1):
        cnt = 0
        for fn in fronts.keys():
            bn = fn + off
            if bn in backs:
                cnt += 1
        if cnt > best_count:
            best_count = cnt
            best_offset = off

    pairs: List[Tuple[Path, Path]] = []
    for fn in sorted(fronts.keys()):
        bn = fn + best_offset
        if bn in backs:
            pairs.append((fronts[fn], backs[bn]))

    return best_offset, pairs


# ----------------------------
# OCR for code on Pro Debut backs
# ----------------------------
CODE_RE = re.compile(r"\b(PDC|PD)[\s\-–_]*0*(\d{1,3})\b", flags=re.IGNORECASE)

def preprocess(img: Image.Image) -> Image.Image:
    img = ImageOps.exif_transpose(img)
    img = img.convert("RGB")
    img = ImageOps.grayscale(img)
    img = ImageEnhance.Contrast(img).enhance(2.2)
    img = ImageEnhance.Sharpness(img).enhance(1.7)
    return img


def roi_boxes(w: int, h: int) -> List[Tuple[str, Tuple[int, int, int, int]]]:
    """
    Multiple ROIs because the code can sit in different places depending on scan/crop.
    For portrait backs, code is often bottom-right corner, but we try a few.
    """
    boxes = []

    # Bottom-right tight corner
    boxes.append(("bottomright_tight", (int(w * 0.70), int(h * 0.78), int(w * 0.98), int(h * 0.98))))
    # Bottom-right wider
    boxes.append(("bottomright_wide", (int(w * 0.60), int(h * 0.70), int(w * 0.98), int(h * 0.98))))
    # Right edge mid-lower (sometimes code sits on a right bar)
    boxes.append(("right_bar", (int(w * 0.78), int(h * 0.35), int(w * 0.98), int(h * 0.98))))
    # Bottom band (in case it’s near bottom center)
    boxes.append(("bottom_band", (int(w * 0.15), int(h * 0.78), int(w * 0.98), int(h * 0.98))))

    return boxes


def threshold_variants(img: Image.Image) -> List[Tuple[str, Image.Image]]:
    """
    Try a few threshold/invert variants to handle light text on dark background.
    """
    variants = [("plain", img)]

    inv = ImageOps.invert(img)
    variants.append(("inv", inv))

    # Simple binary thresholds
    for thr in (140, 165, 185, 205):
        b = img.point(lambda p, t=thr: 255 if p > t else 0)
        variants.append((f"bin{thr}", b))
        bi = ImageOps.invert(b)
        variants.append((f"invbin{thr}", bi))

    return variants


def ocr_image(pimg: Image.Image, psm: int, timeout_s: int) -> str:
    cfg = f"--oem 3 --psm {psm} -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-"
    return pytesseract.image_to_string(pimg, config=cfg, timeout=timeout_s) or ""


def normalize_code(raw: str) -> Optional[str]:
    raw = (raw or "").upper()
    raw = raw.replace(" ", "").replace("_", "-").replace("–", "-")
    # Common OCR confusions
    raw = raw.replace("0D", "PD").replace("P0", "PD").replace("PDC-", "PDC-").replace("PD-", "PD-")

    m = CODE_RE.search(raw)
    if not m:
        return None
    prefix = m.group(1).upper()
    num = int(m.group(2))
    return f"{prefix}-{num}"


def detect_code_from_back(back_path: Path, timeout_s: int, debug_dir: Optional[Path] = None) -> Tuple[Optional[str], str]:
    """
    Returns (code, note). note contains the best attempt details if NO_CODE.
    """
    img0 = Image.open(back_path)
    img0 = ImageOps.exif_transpose(img0)

    # Try rotations: sometimes scans are rotated
    rotations = [0, 90, 180, 270]

    best_note = ""
    start = time.time()

    for rot in rotations:
        if rot != 0:
            imgR = img0.rotate(rot, expand=True)
        else:
            imgR = img0

        img = preprocess(imgR)
        w, h = img.size

        for roi_name, box in roi_boxes(w, h):
            cropped = img.crop(box)
            # upscale
            cropped = cropped.resize((cropped.width * 2, cropped.height * 2))

            for vname, vimg in threshold_variants(cropped):
                # Try a couple PSMs
                for psm in (6, 7, 11):
                    try:
                        raw = ocr_image(vimg, psm=psm, timeout_s=timeout_s)
                    except Exception as e:
                        raw = ""
                    txt = clean_text(raw)
                    code = normalize_code(txt)

                    if code:
                        return code, f"rot={rot},roi={roi_name},v={vname},psm={psm}: {txt}"

                    # update best note occasionally so you can see what it read
                    if txt and len(txt) > len(best_note):
                        best_note = f"rot={rot},roi={roi_name},v={vname},psm={psm}: {txt}"

                    # optional time guard (in addition to pytesseract timeout)
                    if timeout_s and (time.time() - start) > (timeout_s * 10):
                        break

            # Debug save (only if requested)
            if debug_dir:
                debug_dir.mkdir(parents=True, exist_ok=True)
                dbg = debug_dir / f"{back_path.stem}__{rot}__{roi_name}.png"
                try:
                    cropped.save(dbg)
                except Exception:
                    pass

    return None, (best_note or "NO_READ")


# ----------------------------
# Output naming
# ----------------------------
def build_new_name(entry: ChecklistEntry, side: str) -> str:
    """
    Filename format:
      <player_or_team>__<team>__<cardcode>__pro-debut__<side>.jpg

    If Player is blank (teams cards), use Team for both.
    """
    player = entry.player.strip() or entry.team.strip() or "unknown"
    team = entry.team.strip() or entry.player.strip() or "unknown"

    p = slugify(player)
    t = slugify(team)
    c = slugify(entry.code)

    return f"{p}__{t}__{c}__pro-debut__{side}.jpg"


# ----------------------------
# Main
# ----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--front-dir", required=True)
    ap.add_argument("--back-dir", required=True)
    ap.add_argument("--checklist", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--write-back-map", action="store_true")
    ap.add_argument("--timeout", type=int, default=3)
    ap.add_argument("--debug-no-code", action="store_true")
    ap.add_argument("--tesseract", default=r"C:\Program Files\Tesseract-OCR\tesseract.exe")
    args = ap.parse_args()

    pytesseract.pytesseract.tesseract_cmd = args.tesseract

    front_dir = Path(args.front_dir)
    back_dir = Path(args.back_dir)
    checklist = Path(args.checklist)
    out_csv = Path(args.out)

    print("Loading checklist…")
    code_map = load_checklist_map(checklist)
    print(f"Checklist loaded: {len(code_map)} codes")

    offset, pairs = pair_front_back_auto(front_dir, back_dir, max_offset=10)
    print(f"Detected pairing offset: {offset:+d}  (meaning back = front {offset:+d})")
    print(f"Pairs found: {len(pairs)}")

    rows_front = []
    rows_back = []

    ok = 0
    no_code = 0
    no_match = 0

    debug_dir = None
    if args.debug_no_code:
        debug_dir = Path("debug_no_code")

    for i, (front_path, back_path) in enumerate(pairs, start=1):
        print(f"[{i}/{len(pairs)}] OCR code: {back_path.name} …")

        code, note = detect_code_from_back(back_path, timeout_s=args.timeout,
                                           debug_dir=(debug_dir / back_path.stem if debug_dir else None))

        if not code:
            no_code += 1
            rows_front.append({
                "OldName": front_path.name,
                "NewName": "",
                "CardCode": "",
                "Player": "",
                "MinorTeam": "",
                "Status": "NO_CODE",
                "Note": note
            })
            if args.write_back_map:
                rows_back.append({
                    "OldName": back_path.name,
                    "NewName": "",
                    "CardCode": "",
                    "Player": "",
                    "MinorTeam": "",
                    "Status": "NO_CODE",
                    "Note": note
                })
            continue

        entry = code_map.get(code)
        if not entry:
            no_match += 1
            rows_front.append({
                "OldName": front_path.name,
                "NewName": "",
                "CardCode": code,
                "Player": "",
                "MinorTeam": "",
                "Status": "NO_MATCH",
                "Note": note
            })
            if args.write_back_map:
                rows_back.append({
                    "OldName": back_path.name,
                    "NewName": "",
                    "CardCode": code,
                    "Player": "",
                    "MinorTeam": "",
                    "Status": "NO_MATCH",
                    "Note": note
                })
            continue

        ok += 1
        new_front = build_new_name(entry, "front")
        rows_front.append({
            "OldName": front_path.name,
            "NewName": new_front,
            "CardCode": entry.code,
            "Player": entry.player,
            "MinorTeam": entry.team,
            "Status": "OK",
            "Note": ""
        })

        if args.write_back_map:
            new_back = build_new_name(entry, "back")
            rows_back.append({
                "OldName": back_path.name,
                "NewName": new_back,
                "CardCode": entry.code,
                "Player": entry.player,
                "MinorTeam": entry.team,
                "Status": "OK",
                "Note": ""
            })

    out_csv.parent.mkdir(parents=True, exist_ok=True)

    fields = ["OldName", "NewName", "CardCode", "Player", "MinorTeam", "Status", "Note"]
    with out_csv.open("w", newline="", encoding="utf-8") as fp:
        w = csv.DictWriter(fp, fieldnames=fields)
        w.writeheader()
        w.writerows(rows_front)

    print(f"Wrote front rename map: {out_csv.name} (rows={len(rows_front)})")

    if args.write_back_map:
        back_out = out_csv.with_name(out_csv.stem + "_backs" + out_csv.suffix)
        with back_out.open("w", newline="", encoding="utf-8") as fp:
            w = csv.DictWriter(fp, fieldnames=fields)
            w.writeheader()
            w.writerows(rows_back)
        print(f"Wrote back rename map:  {back_out.name} (rows={len(rows_back)})")

    print(f"OK: {ok} | NO_CODE: {no_code} | NO_MATCH: {no_match}")
    if args.debug_no_code:
        print("NO_CODE debug crops saved to: debug_no_code")
    print("Done.")


if __name__ == "__main__":
    main()
