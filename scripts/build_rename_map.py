import argparse
import csv
import re
from pathlib import Path
from difflib import SequenceMatcher

from PIL import Image, ImageOps, ImageEnhance
import pytesseract
import openpyxl


CARDNO_RE = re.compile(r"\b((?:BD|PD|PDC|BDC|BCP|PBP|BPP|B)\s*-\s*\d{1,4})\b", re.IGNORECASE)


def clean_text(s: str) -> str:
    s = (s or "").replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def slugify(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[’'`]", "", s)
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s


def normalize_name(s: str) -> str:
    s = (s or "").strip()
    s = s.replace(",", " ")
    s = re.sub(r"[^A-Za-z \-]", " ", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    s = re.sub(r"\b(bowman|draft|chrome|pro|debut|topps|prospects)\b", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def normalize_team(s: str) -> str:
    s = (s or "").strip()
    s = s.replace(",", " ")
    s = re.sub(r"[^A-Za-z \-]", " ", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    return s


def sim(a: str, b: str) -> float:
    return SequenceMatcher(None, (a or "").lower(), (b or "").lower()).ratio()


def best_match(query: str, candidates: list[str], min_letters: int = 4):
    q = (query or "").strip()
    if len(re.sub(r"[^A-Za-z]", "", q)) < min_letters:
        return ("", 0.0)
    best_c, best_s = "", 0.0
    for c in candidates:
        sc = sim(q, c)
        if sc > best_s:
            best_c, best_s = c, sc
    return best_c, best_s


# ----------------------------
# Checklist dictionary (by Card# AND by Player/Team)
# ----------------------------

def load_checklist_cardmap(xlsx_path: Path, set_token: str):
    """
    Returns:
      cardmap:  { "BD-133": ("Alfredo Duno", "Cincinnati Reds", set_token), ... }
      players:  [ "Alfredo Duno", ... ]
      teams:    [ "Cincinnati Reds", ... ]
    """
    cardmap = {}
    players = set()
    teams = set()

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 3:
                continue

            card, player, team = row[0], row[1], row[2]
            if not isinstance(card, str) or not isinstance(player, str) or not isinstance(team, str):
                continue

            card = card.strip().upper().replace(" ", "")
            # Looks like BD-133 / PD-1 etc.
            if not re.match(r"^[A-Z]{1,5}-\d{1,4}$", card):
                continue

            p = normalize_name(player)
            t = normalize_team(team)
            if not p:
                continue

            players.add(p)
            if t:
                teams.add(t)

            # Prefer first seen; (or you can overwrite—either is fine)
            if card not in cardmap:
                cardmap[card] = (p, t, set_token)

    return cardmap, sorted(players), sorted(teams)


# ----------------------------
# OCR helpers
# ----------------------------

def preprocess(img: Image.Image) -> Image.Image:
    img = ImageOps.exif_transpose(img)
    img = img.convert("RGB")
    img = ImageOps.grayscale(img)
    img = ImageEnhance.Contrast(img).enhance(2.0)
    img = ImageEnhance.Sharpness(img).enhance(1.5)
    return img


def ocr(img: Image.Image, psm: int, whitelist: str | None = None) -> str:
    cfg = f"--oem 3 --psm {psm}"
    if whitelist:
        cfg += f" -c tessedit_char_whitelist={whitelist}"
    return clean_text(pytesseract.image_to_string(img, config=cfg))


def crop_rel(img: Image.Image, x1, y1, x2, y2) -> Image.Image:
    w, h = img.size
    box = (int(w * x1), int(h * y1), int(w * x2), int(h * y2))
    return img.crop(box)


def binarize(img: Image.Image, invert: bool, thresh: int) -> Image.Image:
    img = ImageOps.autocontrast(img)
    if invert:
        img = ImageOps.invert(img)
    img = img.point(lambda p: 255 if p > thresh else 0)
    return img


def find_card_number(img0: Image.Image):
    """
    Try to locate a card number like BD-133 / PD-180.
    We try rotations 0/90/180/270 and OCR a few likely regions.
    Returns (cardno, rotation_degrees, hit_text) or ("", 0, "")
    """
    rotations = [(0, img0), (90, img0.rotate(90, expand=True)), (180, img0.rotate(180, expand=True)), (270, img0.rotate(270, expand=True))]

    for deg, im in rotations:
        im = preprocess(im)

        # Regions that commonly contain card# on backs:
        # - right-middle strip
        # - right-lower strip
        regions = [
            crop_rel(im, 0.72, 0.30, 0.98, 0.72),
            crop_rel(im, 0.70, 0.55, 0.98, 0.95),
            crop_rel(im, 0.60, 0.20, 0.98, 0.50),
        ]

        for r in regions:
            # upscale + binarize helps a lot on small print
            r2 = r.resize((r.width * 3, r.height * 3))
            for inv in (False, True):
                r3 = binarize(r2, invert=inv, thresh=165)
                txt = ocr(r3, psm=6, whitelist="ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz-0123456789")
                m = CARDNO_RE.search(txt)
                if m:
                    cardno = m.group(1).upper().replace(" ", "")
                    return cardno, deg, txt

    return "", 0, ""


def detect_side(img0: Image.Image) -> str:
    """
    Heuristic: if we can find a card number, it's almost certainly a back.
    Otherwise assume front.
    """
    cardno, _, _ = find_card_number(img0)
    return "back" if cardno else "front"


def ocr_bottom_nameplate(img0: Image.Image, fast: bool):
    """
    OCR name + team from the BOTTOM nameplate area (your images match this layout).
    Returns (name_ocr, team_ocr, debug)
    """
    im = preprocess(img0)
    # Bottom bar: name roughly 78-90% height, team roughly 90-97%
    name_crop = crop_rel(im, 0.10, 0.78, 0.90, 0.90)
    team_crop = crop_rel(im, 0.10, 0.90, 0.90, 0.975)

    up = 2 if fast else 3
    name_crop = name_crop.resize((name_crop.width * up, name_crop.height * up))
    team_crop = team_crop.resize((team_crop.width * up, team_crop.height * up))

    # Try both normal and inverted binarization; pick "best-looking" by length
    best_name, best_team = "", ""
    best_dbg = ""

    for inv in (False, True):
        n = binarize(name_crop, invert=inv, thresh=165 if fast else 155)
        t = binarize(team_crop, invert=inv, thresh=170 if fast else 160)

        name_txt = ocr(n, psm=7, whitelist="ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz- ")
        team_txt = ocr(t, psm=6, whitelist="ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz- ")

        name_txt = normalize_name(name_txt)
        team_txt = normalize_team(team_txt)

        score = len(name_txt) + len(team_txt)
        if score > (len(best_name) + len(best_team)):
            best_name, best_team = name_txt, team_txt
            best_dbg = f"invert={inv}"

    return best_name, best_team, best_dbg


def build_filename(player: str, team: str, set_token: str, side: str) -> str:
    p = slugify(player)[:40] or "unknown-player"
    t = slugify(team)[:35] if team else ""
    st = slugify(set_token)[:40] if set_token else "unknown-set"
    parts = [p]
    if t:
        parts.append(t)
    parts.append(st)
    parts.append(side)  # front/back
    return "__".join(parts) + ".jpg"


# ----------------------------
# Main
# ----------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True, help="Folder containing images")
    ap.add_argument("--out", required=True, help="Output CSV path (rename map)")
    ap.add_argument("--tesseract", default=r"C:\Program Files\Tesseract-OCR\tesseract.exe")

    ap.add_argument("--checklist-prodebut", required=True, help="Pro Debut checklist xlsx path")
    ap.add_argument("--checklist-draft", required=True, help="Bowman Draft checklist xlsx path")
    ap.add_argument("--token-prodebut", default="topps-pro-debut-2025")
    ap.add_argument("--token-draft", default="bowman-draft-2025")

    ap.add_argument("--min-player-score", type=float, default=0.84)
    ap.add_argument("--min-team-score", type=float, default=0.78)
    ap.add_argument("--fast", action="store_true")
    ap.add_argument("--max-files", type=int, default=0)

    args = ap.parse_args()
    pytesseract.pytesseract.tesseract_cmd = args.tesseract

    # Load dictionaries (card# maps + name lists)
    pro_cardmap, pro_players, pro_teams = load_checklist_cardmap(Path(args.checklist_prodebut), args.token_prodebut)
    dr_cardmap, dr_players, dr_teams = load_checklist_cardmap(Path(args.checklist_draft), args.token_draft)

    # Merge card maps (card# disambiguates set)
    all_cardmap = {}
    all_cardmap.update(pro_cardmap)
    all_cardmap.update(dr_cardmap)

    print(f"Loaded Pro Debut: {len(pro_cardmap)} cards, {len(pro_players)} players")
    print(f"Loaded Draft:    {len(dr_cardmap)} cards, {len(dr_players)} players")
    print(f"Total card# map: {len(all_cardmap)}")

    src_dir = Path(args.dir)
    files = sorted([p for p in src_dir.iterdir() if p.is_file() and p.suffix.lower() in [".jpg", ".jpeg", ".png", ".webp"]])
    if args.max_files and args.max_files > 0:
        files = files[:args.max_files]

    rows = []
    total = len(files)

    for idx, f in enumerate(files, start=1):
        print(f"[{idx}/{total}] {f.name}")
        try:
            img0 = Image.open(f)

            # 1) Card number lookup (most accurate)
            cardno, rot, hit = find_card_number(img0)
            side = "back" if cardno else "front"

            if cardno and cardno in all_cardmap:
                player, team, set_token = all_cardmap[cardno]
                new_name = build_filename(player, team, set_token, side)
                note = "cardno->checklist exact"
                rows.append({
                    "OldName": f.name,
                    "NewName": new_name,
                    "Side": side,
                    "CardNumber": cardno,
                    "SetToken": set_token,
                    "Player": player,
                    "Team": team,
                    "PlayerScore": 1.0,
                    "TeamScore": 1.0,
                    "Rotation": rot,
                    "Note": note,
                    "Error": ""
                })
                continue

            # 2) Otherwise OCR bottom nameplate and fuzzy-match BOTH sets
            player_ocr, team_ocr, dbg = ocr_bottom_nameplate(img0, fast=args.fast)

            pro_p, pro_ps = best_match(player_ocr, pro_players)
            dr_p,  dr_ps  = best_match(player_ocr, dr_players)

            # choose the set with better player score
            if pro_ps >= dr_ps:
                set_token = args.token_prodebut
                chosen_players, chosen_teams = pro_players, pro_teams
                player_match, player_score = pro_p, pro_ps
            else:
                set_token = args.token_draft
                chosen_players, chosen_teams = dr_players, dr_teams
                player_match, player_score = dr_p, dr_ps

            team_match, team_score = ("", 0.0)
            if team_ocr:
                team_match, team_score = best_match(team_ocr, chosen_teams, min_letters=3)

            if player_score >= args.min_player_score:
                final_team = team_match if team_score >= args.min_team_score else ""
                new_name = build_filename(player_match, final_team, set_token, side)
                note = "nameplate->fuzzy"
            else:
                new_name = ""
                note = "low-confidence (blank NewName)"

            rows.append({
                "OldName": f.name,
                "NewName": new_name,
                "Side": side,
                "CardNumber": cardno,
                "SetToken": set_token,
                "Player": player_match,
                "Team": team_match,
                "PlayerScore": round(player_score, 3),
                "TeamScore": round(team_score, 3),
                "Rotation": rot,
                "Note": f"{note}; dbg={dbg}; ocr_player='{player_ocr}'; ocr_team='{team_ocr}'",
                "Error": ""
            })

        except Exception as e:
            rows.append({
                "OldName": f.name,
                "NewName": "",
                "Side": "",
                "CardNumber": "",
                "SetToken": "",
                "Player": "",
                "Team": "",
                "PlayerScore": 0,
                "TeamScore": 0,
                "Rotation": 0,
                "Note": "",
                "Error": str(e)
            })

    out_csv = Path(args.out)
    out_csv.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = ["OldName","NewName","Side","CardNumber","SetToken","Player","Team","PlayerScore","TeamScore","Rotation","Note","Error"]
    with out_csv.open("w", newline="", encoding="utf-8") as fp:
        w = csv.DictWriter(fp, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)

    print(f"\nWrote: {out_csv}")
    print("Tip: rows with blank NewName are safe (PowerShell will SKIP).")


if __name__ == "__main__":
    main()
